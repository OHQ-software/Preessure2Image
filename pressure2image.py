# 面圧センサデータから圧力分布画像データへ変換
import os			# ディレクトリアクセス
import csv			# CSVアクセス
import openpyxl		# エクセルファイル制御
from openpyxl.styles import PatternFill		# セルの色を変更する関数

number_data_pressure = 256					# 圧力データ数/行
side_length_sensor = 16						# 圧力センサの1辺のデータ数 number_data_pressureの平方根

def get_file_list(dir_l, dir_r, extension):
	"""
	Summary:
		解析対象ファイルペアリストを取得
	Args:
		(str)dir_l: 探索対象フォルダ名1
		(str)dir_r: 探索対象フォルダ名2
		(str)extension: 対象ファイル拡張子(ドット無し)
	Returns:
		([str])file_list: ファイルリスト
	"""
	files_l = []
	files_r = []
	file_list = []
	files_l = os.listdir(dir_l)									# ファイル取得
	files_r = os.listdir(dir_r)
	file_list = set(files_l) & set(files_r)						# 同じファイル名のリスト
	file_list = [f for f in file_list if '.'+extension in f]	# 指定拡張子のみを抽出
	file_list.sort()											# 破壊的ソート
	return file_list

def get_pressure_list(file_puressure):
	"""
	Summary:
		圧力データリスト(圧力,時間)を取得
	Args:
		(str)file_puressure : 圧力データファイル名
	Returns:
		([float])press_list: 圧力データ一次元配列
		([float])time_list: 時間データ一次元配列
	"""
	press_list = []
	time_list = []
	with open(file_puressure, 'r') as file:
		header = next(csv.reader(file))		# ヘッダー行を空読み
		reader = csv.reader(file)			# データ行を読出し
		for line in reader:					# データ→リスト
			# line[0]=番号, line[1]=時間[sec], line[2]=カフ圧[mmHg], その他情報....
			press_list.append(float(line[2]))
			time_list.append(float(line[1]))
	return press_list, time_list

def get_max_pressure(press_list, time_list):
	"""
	Summary:
		最高加圧情報(圧力,時間)を取得
		前後の圧力差が1以上であればノイズと判定する
	Args:
		([float])press_list: 圧力データリスト
		([float])time_list: 圧力データリスト(要素番号がpress_listと同じ)
	Returns:
		(float)max_pres: 検出最大圧力
		(float)max_pres_time: 最大圧力検出時の時間
	"""
	max_pres = 0.0
	max_pres_time = 0.0
	for i in range(1, len(press_list)-1):				# リストデータから最大値を取得
		if abs(press_list[i-1] - press_list[i]) < 1.0 and (press_list[i] - press_list[i+1]) < 1.0:	# ノイズ判定
			if max_pres <= press_list[i]:				# 最大値判定
				max_pres = press_list[i]
				max_pres_time = time_list[i]
	return max_pres, max_pres_time

def get_time_reduced_specified_pressure(pressure_list, time_list, spec_pres, max_pres):
	"""
	Summary:
		最高加圧達成後に規定圧力へ減圧するまでの時間を取得
		前後の圧力差が1以上であればノイズと判定する
	Args:
		([float])pressure_list: 圧力データリスト
		([float])time_list: 時間リスト
		(float)spec_pres: 規定圧力
		(float)max_pres: 最高圧
	Returns:
		(float)time_reduced_specified_pressure: 最高加圧達成後に規定圧力へ減圧するまでの時間
	"""
	flg_reached_max_pressure = False
	time_reduced_specified_pressure = 0.0
	for i in range(1, len(pressure_list)-1):
		if abs(pressure_list[i-1] - pressure_list[i]) < 1.0 and abs(pressure_list[i] - pressure_list[i+1]) < 1.0:	# ノイズ判定
			if pressure_list[i] >= max_pres:												# 最大圧力超え判定
				flg_reached_max_pressure = True
			if flg_reached_max_pressure == True and pressure_list[i] <= spec_pres:			# 最大加圧超えた後に規定圧力まで減圧したか
				time_reduced_specified_pressure = time_list[i]
				break				# 取得できたので検索ループを脱出
	return time_reduced_specified_pressure

def get_pressure_list_surface_sensor(file_sensor):
	"""
	Summary:
		面圧センサデータの時間毎の平均圧力値を取得
		データのヘッダ行は5行としている
		センサデータ数は、256個/行としている
	Args:
		(str)file_sensor: 面圧センサデータファイル名
	Returns:
		([float])avg_pres_list: 各時間での圧力平均値データリスト
		([float])time_list: 時間リスト(要素番号はavg_pres_listと同じ)
	"""
	avg_pres_list = []
	time_list = []
	sum_list = []
	header_lines = 5							# ヘッダー行数
	with open(file_sensor, 'r') as file:
		for i in range(header_lines):
			header = next(csv.reader(file))		# ヘッダー行を空読み
		reader = csv.reader(file)				# データ行を読出し
		for line in reader:						# データ→リスト
			# line[0]=時間, line[1]～line[256]:Elem0～Elem255の圧力
			total_press = 0
			for i in range(1,number_data_pressure+1):	# line[1](=Elem0)～
				total_press = total_press + float(line[i]) * 100		# floatのままだと小数点演算で誤差が出るので整数にする
			avg_pres_list.append(total_press/number_data_pressure/100)
			time_list.append(float(line[0]))
	return avg_pres_list, time_list

def get_time_max_pressure_surface_sensor(avg_sensor_pres_list, sensor_time_list):
	"""
	Summary:
		面圧データより、最高加圧となった時間tmaxを取得
		前後の圧力差が1以上であればノイズと判定する
	Args:
		([float])avg_pres_list: 各時間での圧力平均値データリスト
		([float])time_list: 時間リスト(要素番号はavg_pres_listと同じ)
	Returns:
		(float)time_max_pressure_surface_sensor: 最高加圧となった時間
	"""
	time_max_pressure_surface_sensor = 0.0
	max_pres = 0
	for i in range(1, len(avg_sensor_pres_list)-1):
		if abs(avg_sensor_pres_list[i-1] - avg_sensor_pres_list[i]) < 1.0 and abs(avg_sensor_pres_list[i] - avg_sensor_pres_list[i+1]) < 1.0:
			if max_pres <= avg_sensor_pres_list[i]:
				max_pres = avg_sensor_pres_list[i]
				time_max_pressure_surface_sensor = sensor_time_list[i]
	return time_max_pressure_surface_sensor

def get_surface_sensor_data_target_press(file_sensor, time_max_pressure_surface_sensor, time_specified_pressure):
	"""
	Summary:
		tmaxからTp経過後の面圧センサデータを取得
	Args:
		(str)file_sensor: 面圧センサデータファイル名
		time_max_pressure_surface_sensor:
		time_specified_pressure:
	Returns:
		([float])surface_sensor_data_target_press: 面圧センサデータ(number_data_pressure個)
	"""
	surface_sensor_data_target_press = []
	with open(file_sensor, 'r') as file:
		for i in range(5):
			header = next(csv.reader(file))		# ヘッダー行を空読み
		reader = csv.reader(file)				# データ行を読出し
		for line in reader:						# データ→リスト
			# line[0]=時間, line[1]～line[256]:Elen0～Elem255の圧力
			if float(line[0]) >= time_max_pressure_surface_sensor + time_specified_pressure:
				print("面圧センサデータが指定圧力まで減圧した事を検出した時間:" + line[0])
				for i in range(1,number_data_pressure+1):
					surface_sensor_data_target_press.append(float(line[i]))
				break							# 検出したのでループを脱出
	return surface_sensor_data_target_press

def convert_linear2matrix(array_linear):
	"""
	Summary:
		一次元配列を二次元配列へ変換
		展開するときの行列番号に注意(転置してるかもしれない)
	Args:
		([])array_linear: 一次元配列
	Returns:
		([[]])matrix: 二次元配列
	"""
	matrix = []
	for i in range(side_length_sensor):
		line = []
		for j in range(side_length_sensor-1, -1, -1):
			line.append(array_linear[i*side_length_sensor+j])
		matrix.append(line)
	return matrix

def smooth_matrix(matrix_in, size):
	"""
	Summary:
		二次元配列の平滑化
		近傍(size*size)分の平均値を対象点の値とする平滑化を行う
		sizeが偶数の場合は考慮してない
	Args:
		([[float]])matrix_in: 平滑化対象二次元配列
		(int)size: 平均する対象範囲
	Returns:
		([[float]])matrix_out: 平滑化後二次元配列
	"""
	matrix_out = [[0 for x in range(side_length_sensor)] for y in range(side_length_sensor)]
	for y in range(int(0+(size-1)/2),int(side_length_sensor-(size-1)/2)):			# 対象点matrix[y][x]ループ
		for x in range(int(0+(size-1)/2),int(side_length_sensor-(size-1)/2)):
			for j in range(int(-(size-1)/2), int((size-1)/2+1)):					# 近傍点ループ
				for i in range(int(-(size-1)/2), int((size-1)/2+1)):
					matrix_out[y][x] += matrix_in[y+j][x+i]
			matrix_out[y][x] = round(matrix_out[y][x] / (size*size), 2)				# 小数点第3位以下四捨五入(銀行丸め)
	return matrix_out

def make_excel_sheet(matrix, workbook, sheet_name):
	"""
	Summary:
		圧力に応じて色づけしたシートを作成し、エクセルブックに追加していく
			シートは新規追加していくので,デフォルトの"Sheet"という空シートは残る
			0以下は白にする 値の無い点も0になるため
			300以上はcolor_codeの一番濃い色にする
	Args:
		([[float]])matrix: 圧力分布データ二次元配列
		(str)workbook: エクセルブック
		(str)sheet_name: 追加するシート名前(元は解析対象ファイル*.csv)
	"""
	color_code = [	'7030A0',	# 紫       0～
					'002060',	# 濃い青  30～
					'0070C0',	# 青      60～
					'00B0F0',	# 水色    90～
					'00B050',	# 緑     120～
					'92D050',	# 薄い緑 150～
					'FFFF00',	# 黄     180～
					'FFC000',	# 橙     210～
					'FF0000',	# 赤     240～
					'C00000' ]	# 濃い赤 270～
	num_sheet = len(workbook.sheetnames)							# ブックのシート数取得 シート番号作成のため
	workbook.create_sheet(index=num_sheet+1, title=sheet_name)		# シート作成
	ws = workbook.worksheets[num_sheet]								# 作成したシート(最後のシート番号)を指定
	for j in range(side_length_sensor):
		for i in range(side_length_sensor):
			ws.cell(i+1,j+1).value = matrix[j][i]					# セルに値を代入
			if matrix[j][i] >= 300:									# 値に応じて色を決定
				fgColor = color_code[len(color_code)-1]
			elif matrix[j][i] <= 0:
				fgColor = 'FFFFFF'
			else:
				fgColor = color_code[int(matrix[j][i]/30)]
			ws.cell(i+1,j+1).fill = PatternFill(patternType='solid', fgColor=fgColor)	# セルを決定した色にする

def get_average_matrix_list(matrix_list):
	"""
	Summary:
		全平滑化データの圧力を平均した圧力分布データ作成
	Args:
		([[[float]])matrix_list: 圧力分布データ二次元配列のリスト
	Returns:
		([[float]])matrix_avg: 平均した圧力分布データ二次元配列
	"""
	matrix_avg = [[0 for x in range(side_length_sensor)] for y in range(side_length_sensor)]	# 初期値0の二次元配列を作成
	for matrix in matrix_list:																	# 全圧力分布データの同じ位置の値を加算する
		for j in range(side_length_sensor):
			for i in range(side_length_sensor):
				matrix_avg[j][i] = matrix_avg[j][i] + matrix[j][i]
	for j in range(side_length_sensor):															# 加算した結果を圧力分布データ数で割って平均化する
		for i in range(side_length_sensor):
			matrix_avg[j][i] = round(matrix_avg[j][i] / len(matrix_list), 2)
	return matrix_avg

def main():
	""" メインルーチン """
	print('面圧センサデータから圧力分布データファイルを作成')

	# ユーザー設定値入力
	print('検出圧力:')
	target_press = float(input())
	print('圧力平均するサイズ(NxNでNは奇数):')
	average_size = int(input())

	current_directory = os.getcwd()								# カレントディレクトリを取得
	dir_pressure = current_directory + '/' + "EG1データ"		# EG1データ保存フォルダ
	dir_sensor   = current_directory + '/' + "面圧データ"		# 面圧データ保存フォルダ
	workbook = openpyxl.Workbook()								# 出力用エクセルファイル
	smooth_matrix_list = []										# 平滑化した圧力分布データのリスト

	# 解析対象ファイルペアを取得
	file_list = get_file_list(dir_pressure, dir_sensor, "csv")
	print('対象ファイル:', file_list)

	# ファイル数分処理する
	for file in file_list:
		print('----------------------------------------')
		print('解析ファイル:', file)
		# EG1データより、最高加圧までの時間Tmaxを取得
		pressure_list, time_list = get_pressure_list(dir_pressure + '/' + file)
		max_pressure, max_pressure_time = get_max_pressure(pressure_list, time_list)
		print("最大圧力:", max_pressure, ",  時間:", max_pressure_time)

		# 最大加圧後指定圧力までの減圧時間Tp(≡time_specified_pressure)を取得
		time_specified_pressure = get_time_reduced_specified_pressure(pressure_list, time_list, target_press, max_pressure) - max_pressure_time
		print("指定圧力まで減圧するまでの時間:", time_specified_pressure)

		# 面圧センサデータの時間毎の平均圧力値を取得
		avg_sensor_pres_list, sensor_time_list = get_pressure_list_surface_sensor(dir_sensor + '/' + file)

		# 面圧センサデータより、最高加圧となった時間tmaxを取得
		time_max_pressure_surface_sensor = get_time_max_pressure_surface_sensor(avg_sensor_pres_list, sensor_time_list)
		print("面圧センサが検出した最大圧力に達した時間:", time_max_pressure_surface_sensor)

		# tmaxからTp経過後の面圧センサデータを取得
		surface_sensor_data_target_press = get_surface_sensor_data_target_press(dir_sensor + '/' + file, time_max_pressure_surface_sensor, time_specified_pressure)
		# print(surface_sensor_data_target_press)

		# 面圧センサデータ(一次元配列)を二次元配列へ展開する
		matrix = convert_linear2matrix(surface_sensor_data_target_press)
		# print(matrix)

		# 面圧センサデータの各点のデータを、近傍の点より平均化する(3x3,5x5など)
		smoothed_matrix = smooth_matrix(matrix, average_size)
		# print(smoothed_matrix)
		smooth_matrix_list.append(smoothed_matrix)

		# 圧力に応じて色づけしたエクセルファイルを作成する
		make_excel_sheet(smoothed_matrix, workbook, file.replace('.csv',''))

	# 全平滑化データの圧力を平均した圧力分布データ作成してエクセルファイルへ追加
	average_matrix_list = get_average_matrix_list(smooth_matrix_list)
	make_excel_sheet(average_matrix_list, workbook, '平均')

	# エクセルファイル保存
	workbook.save(current_directory + '/output.xlsx')

	print("終わり")
	owari = input()

main()
